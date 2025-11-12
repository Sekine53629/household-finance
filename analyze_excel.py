"""
Excelファイルの構造を解析して、Djangoモデル設計に必要な情報を抽出
"""
import openpyxl
from openpyxl.utils import get_column_letter
from collections import defaultdict
import re


def analyze_excel(file_path):
    """Excelファイルを解析して構造を出力"""

    try:
        # スタイル情報は無視して読み込み
        wb = openpyxl.load_workbook(file_path, data_only=False, keep_vba=False, rich_text=True)
        print("=" * 80)
        print(f"ファイル: {file_path}")
        print("=" * 80)
        print()

        # シート一覧
        print("【シート一覧】")
        for idx, sheet_name in enumerate(wb.sheetnames, 1):
            print(f"{idx}. {sheet_name}")
        print()

        # 各シートの詳細分析
        for sheet_name in wb.sheetnames:
            print("\n" + "=" * 80)
            print(f"シート: {sheet_name}")
            print("=" * 80)

            ws = wb[sheet_name]

            # シートのサイズ
            print(f"\n最大行: {ws.max_row}, 最大列: {ws.max_column}")

            # ヘッダー行を探す（最初の10行をチェック）
            print("\n【先頭データ（最初の10行×10列）】")
            for row_idx in range(1, min(11, ws.max_row + 1)):
                row_data = []
                for col_idx in range(1, min(11, ws.max_column + 1)):
                    cell = ws.cell(row_idx, col_idx)
                    value = cell.value
                    if value:
                        row_data.append(f"{get_column_letter(col_idx)}{row_idx}: {value}")

                if row_data:
                    print(f"行{row_idx}: {', '.join(row_data[:5])}")  # 最初の5列のみ表示

            # 数式を含むセルを抽出
            print(f"\n【数式を含むセル（最初の30個）】")
            formula_count = 0
            formulas_by_pattern = defaultdict(list)

            for row in ws.iter_rows(min_row=1, max_row=min(100, ws.max_row)):
                for cell in row:
                    if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                        formula = cell.value
                        # 数式のパターンを簡略化
                        pattern = re.sub(r'\d+', 'N', formula)  # 数字をNに置換
                        pattern = re.sub(r'"[^"]*"', '"..."', pattern)  # 文字列を...に置換

                        formulas_by_pattern[pattern].append(f"{cell.coordinate}")

                        if formula_count < 30:
                            coord = cell.coordinate
                            print(f"{coord}: {formula[:100]}{'...' if len(formula) > 100 else ''}")
                            formula_count += 1

            # パターン別の数式集計
            if formulas_by_pattern:
                print(f"\n【数式パターンの集計】")
                for idx, (pattern, cells) in enumerate(sorted(formulas_by_pattern.items(), key=lambda x: len(x[1]), reverse=True)[:10], 1):
                    print(f"{idx}. 使用箇所: {len(cells)}箇所")
                    print(f"   パターン: {pattern[:100]}{'...' if len(pattern) > 100 else ''}")
                    print(f"   セル例: {', '.join(cells[:5])}")
                    print()

            # 名前付き範囲
            print(f"\n【名前付き範囲】")
            named_ranges = []
            for name in wb.defined_names.definedName:
                if name.value:
                    named_ranges.append(f"{name.name}: {name.value}")

            if named_ranges:
                for nr in named_ranges[:20]:  # 最初の20個
                    print(f"  {nr}")
            else:
                print("  なし")

        # 全体のサマリー
        print("\n\n" + "=" * 80)
        print("【分析サマリー】")
        print("=" * 80)
        print(f"シート数: {len(wb.sheetnames)}")

        total_formulas = 0
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            sheet_formulas = sum(1 for row in ws.iter_rows() for cell in row
                               if cell.value and isinstance(cell.value, str) and cell.value.startswith('='))
            total_formulas += sheet_formulas
            print(f"  {sheet_name}: {sheet_formulas}個の数式")

        print(f"\n合計数式数: {total_formulas}")

    except Exception as e:
        print(f"エラーが発生しました: {e}")
        import traceback
        traceback.print_exc()


if __name__ == "__main__":
    file_path = r"C:\Users\imao3\Downloads\給与計算用エクセル202510.xlsm"
    analyze_excel(file_path)
